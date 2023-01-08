from instagram_graph_api import InstagramGraphAPI

import openpyxl

BUSINESS_ACCOUNT_ID='個人でセット'
USER_NAME='個人でセット'
ACCESS_TOKEN='個人でセット'

def convert_result_to_excel(result,excel_file_path):
  wb=openpyxl.Workbook()
  sheet=wb['Sheet']

  header_list=list(result[0].keys())
  # print(header_list)
  for c in range(len(header_list)):
    sheet.cell(row=1,column=c+1).value=header_list[c]
    # print(header_list[c])

  for r in range(len(result)):
    for c in range(len(header_list)):
      sheet.cell(row=r+2,column=c+1).value=result[r][header_list[c]]
      # print(result[r][header_list[c]])

  wb.save(excel_file_path)

def main():
  try:
    instagram_graph_api=InstagramGraphAPI()
    # ex)data_fields='timestamp,like_count,permalink'
    data_fields='個人でセット'
    result=instagram_graph_api.get_media_json(BUSINESS_ACCOUNT_ID,USER_NAME,ACCESS_TOKEN,data_fields)
    # print(result)
    output_path='個人でセット'

    convert_result_to_excel(result,output_path)

  except Exception as e:
    print(e)
    print(type(e))
    print(type(e))

if __name__ == "__main__":
  main()